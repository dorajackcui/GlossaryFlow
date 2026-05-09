#!/usr/bin/env python3
"""Workflow for splitting, pre-filling, checking, and filling term Excel files."""

from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


STEM_SHEET = "主干术语表"
SUFFIX_SHEET = "后缀术语表"
SUMMARY_SHEET = "汇总术语表"
SOURCE_SUFFIX_MARKER = "·"
SOURCE_SUFFIX_MARKERS = (SOURCE_SUFFIX_MARKER, "-")
TARGET_SUFFIX_MARKERS = (":", "：")
TARGET_HYPHEN_SUFFIX_MARKER = "-"
TERM_SHEETS = (STEM_SHEET, SUFFIX_SHEET)


@dataclass(frozen=True)
class TermRow:
    sheet_name: str
    row_index: int
    term_type: str
    source: str
    target: str


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def clean_target_cell(value: object) -> str:
    text = "" if value is None else str(value).replace("\u00a0", " ")
    stripped = text.strip()
    if stripped.startswith((":", "：")):
        return f" {stripped}"
    return stripped


def normalize_header(value: object) -> str:
    return clean_cell(value).lower().replace(" ", "").replace("_", "")


def find_header_column(header: list[str], names: set[str]) -> int | None:
    return next((index for index, name in enumerate(header) if name in names), None)


def split_source(value: str) -> tuple[str, str]:
    marker_positions = [
        (value.index(marker), marker) for marker in SOURCE_SUFFIX_MARKERS if marker in value
    ]
    if not marker_positions:
        return value.strip(), ""

    marker_index, marker = min(marker_positions, key=lambda item: item[0])
    stem = value[:marker_index]
    suffix = value[marker_index + len(marker) :]
    return stem.strip(), f"{marker}{suffix.strip()}"


def find_target_suffix_start(value: str, allow_hyphen_suffix: bool) -> int | None:
    marker_positions = [
        value.index(marker) for marker in TARGET_SUFFIX_MARKERS if marker in value
    ]
    if marker_positions:
        marker_index = min(marker_positions)
    elif allow_hyphen_suffix and TARGET_HYPHEN_SUFFIX_MARKER in value:
        marker_index = value.index(TARGET_HYPHEN_SUFFIX_MARKER)
    else:
        return None

    suffix_start = marker_index
    while suffix_start > 0 and value[suffix_start - 1].isspace():
        suffix_start -= 1
    return suffix_start


def split_target(value: str, allow_hyphen_suffix: bool = False) -> tuple[str, str]:
    if not value:
        return "", ""

    suffix_start = find_target_suffix_start(value, allow_hyphen_suffix)
    if suffix_start is None:
        return value.strip(), ""

    return value[:suffix_start].strip(), value[suffix_start:]


def first_value_wins_add(glossary: dict[str, str], source: str, target: str = "") -> None:
    if source and source not in glossary:
        glossary[source] = target


def first_non_empty_value_wins_add(glossary: dict[str, str], source: str, target: str = "") -> None:
    if not source:
        return
    if source not in glossary or (not glossary[source] and target):
        glossary[source] = target


def find_header_columns(rows: list[tuple[object, ...]]) -> tuple[int, int, int | None]:
    if not rows:
        raise ValueError("Input sheet is empty.")

    header = [normalize_header(cell) for cell in rows[0]]
    source_names = {"source", "src", "源文", "原文", "源语言", "源术语"}
    target_names = {"target", "tgt", "译文", "目标", "目标语言", "目标术语"}

    source_col = next((index for index, name in enumerate(header) if name in source_names), None)
    target_col = next((index for index, name in enumerate(header) if name in target_names), None)
    if source_col is not None:
        return 1, source_col, target_col

    return 0, 0, 1 if len(rows[0]) > 1 else None


def find_summary_header_columns(rows: list[tuple[object, ...]]) -> tuple[int, int, int, int | None]:
    if not rows:
        raise ValueError("Input sheet is empty.")

    header = [normalize_header(cell) for cell in rows[0]]
    term_type_names = {"termtype", "类型", "术语类型"}
    source_names = {"source", "src", "源文", "原文", "源语言", "源术语"}
    target_names = {"target", "tgt", "译文", "目标", "目标语言", "目标术语"}

    term_type_col = find_header_column(header, term_type_names)
    source_col = find_header_column(header, source_names)
    target_col = find_header_column(header, target_names)
    if term_type_col is None or source_col is None:
        raise ValueError(f"{SUMMARY_SHEET} 需要包含 term_type 和 source 列。")
    return 1, term_type_col, source_col, target_col


def find_sheet_with_source(workbook) -> str:
    for sheet in workbook.worksheets:
        first_row = list(sheet.iter_rows(values_only=True, max_row=1))
        if not first_row:
            continue
        headers = {normalize_header(cell) for cell in first_row[0]}
        if headers & {"source", "src", "源文", "原文", "源语言", "源术语"}:
            return sheet.title
    return workbook.worksheets[0].title


def read_source_target_rows(path: Path, sheet_name: str | None = None) -> tuple[str, list[tuple[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        actual_sheet = sheet_name or find_sheet_with_source(workbook)
        if actual_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet {actual_sheet!r} not found in {path}.")

        rows = list(workbook[actual_sheet].iter_rows(values_only=True))
        start_row, source_col, target_col = find_header_columns(rows)
        result: list[tuple[str, str]] = []
        for row in rows[start_row:]:
            source = clean_cell(row[source_col] if source_col < len(row) else "")
            target = clean_cell(row[target_col] if target_col is not None and target_col < len(row) else "")
            if source:
                result.append((source, target))
        return actual_sheet, result
    finally:
        workbook.close()


def split_terms_from_rows(rows: list[tuple[str, str]]) -> tuple[dict[str, str], dict[str, str]]:
    stems: dict[str, str] = {}
    suffixes: dict[str, str] = {}

    for source, target in rows:
        source_stem, source_suffix = split_source(source)
        target_stem, target_suffix = split_target(
            target, allow_hyphen_suffix=source_suffix.startswith(TARGET_HYPHEN_SUFFIX_MARKER)
        )

        first_non_empty_value_wins_add(stems, source_stem, target_stem)
        if source_suffix:
            first_non_empty_value_wins_add(suffixes, source_suffix, target_suffix)

    return stems, suffixes


def write_terms_workbook(
    output_path: Path,
    stems: dict[str, str],
    suffixes: dict[str, str],
    include_status: bool = True,
) -> None:
    workbook = Workbook()
    stem_sheet = workbook.active
    stem_sheet.title = STEM_SHEET
    suffix_sheet = workbook.create_sheet(SUFFIX_SHEET)

    for sheet, glossary in ((stem_sheet, stems), (suffix_sheet, suffixes)):
        sheet.append(["source", "target", "status"] if include_status else ["source", "target"])
        for source, target in glossary.items():
            if include_status:
                sheet.append([source, target or None, "matched" if target else "needs_translation"])
            else:
                sheet.append([source, target or None])
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 42
        if include_status:
            sheet.column_dimensions["C"].width = 20

    workbook.save(output_path)


def write_summary_terms_workbook(
    output_path: Path,
    stems: dict[str, str],
    suffixes: dict[str, str],
    include_status: bool = True,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SUMMARY_SHEET
    sheet.append(["term_type", "source", "target", "status"] if include_status else ["term_type", "source", "target"])

    for sheet_name, glossary in ((STEM_SHEET, stems), (SUFFIX_SHEET, suffixes)):
        for source, target in glossary.items():
            if include_status:
                sheet.append([sheet_name, source, target or None, "matched" if target else "needs_translation"])
            else:
                sheet.append([sheet_name, source, target or None])

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 32
    sheet.column_dimensions["C"].width = 42
    if include_status:
        sheet.column_dimensions["D"].width = 20
    workbook.save(output_path)


def split_demand_workbook(
    input_path: Path,
    output_path: Path,
    sheet_name: str | None = None,
    summary_output_path: Path | None = None,
) -> None:
    _, rows = read_source_target_rows(input_path, sheet_name)
    stems, suffixes = split_terms_from_rows(rows)
    write_terms_workbook(output_path, stems, suffixes, include_status=True)
    if summary_output_path is not None:
        write_summary_terms_workbook(summary_output_path, stems, suffixes, include_status=True)


def read_terms_workbook(path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SUMMARY_SHEET in workbook.sheetnames:
            return read_summary_terms_workbook(workbook)

        terms: dict[str, dict[str, str]] = {}

        for sheet_name in TERM_SHEETS:
            if sheet_name not in workbook.sheetnames:
                terms[sheet_name] = {}
                continue

            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            if not rows:
                terms[sheet_name] = {}
                continue

            start_row, source_col, target_col = find_header_columns(rows)
            sheet_terms: dict[str, str] = {}
            for row in rows[start_row:]:
                source = clean_cell(row[source_col] if source_col < len(row) else "")
                target = clean_target_cell(row[target_col] if target_col is not None and target_col < len(row) else "")
                first_value_wins_add(sheet_terms, source, target)
            terms[sheet_name] = sheet_terms

        return terms
    finally:
        workbook.close()


def read_summary_terms_workbook(workbook) -> dict[str, dict[str, str]]:
    terms: dict[str, dict[str, str]] = {sheet_name: {} for sheet_name in TERM_SHEETS}
    rows = list(workbook[SUMMARY_SHEET].iter_rows(values_only=True))
    if not rows:
        return terms

    start_row, term_type_col, source_col, target_col = find_summary_header_columns(rows)
    for row in rows[start_row:]:
        term_type = clean_cell(row[term_type_col] if term_type_col < len(row) else "")
        source = clean_cell(row[source_col] if source_col < len(row) else "")
        target = clean_target_cell(row[target_col] if target_col is not None and target_col < len(row) else "")
        if not source:
            continue
        if term_type not in terms:
            raise ValueError(f"{SUMMARY_SHEET} 的 term_type 需要是 {STEM_SHEET} 或 {SUFFIX_SHEET}：{term_type}")
        first_value_wins_add(terms[term_type], source, target)
    return terms


def read_terms_entries(path: Path) -> dict[str, list[tuple[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SUMMARY_SHEET in workbook.sheetnames:
            return read_summary_terms_entries(workbook)

        entries: dict[str, list[tuple[str, str]]] = {}

        for sheet_name in TERM_SHEETS:
            if sheet_name not in workbook.sheetnames:
                entries[sheet_name] = []
                continue

            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            if not rows:
                entries[sheet_name] = []
                continue

            start_row, source_col, target_col = find_header_columns(rows)
            sheet_entries: list[tuple[str, str]] = []
            for row in rows[start_row:]:
                source = clean_cell(row[source_col] if source_col < len(row) else "")
                target = clean_target_cell(row[target_col] if target_col is not None and target_col < len(row) else "")
                if source:
                    sheet_entries.append((source, target))
            entries[sheet_name] = sheet_entries

        return entries
    finally:
        workbook.close()


def read_term_rows(path: Path) -> list[TermRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SUMMARY_SHEET in workbook.sheetnames:
            return read_summary_term_rows(workbook)

        term_rows: list[TermRow] = []
        for sheet_name in TERM_SHEETS:
            if sheet_name not in workbook.sheetnames:
                continue

            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            if not rows:
                continue

            start_row, source_col, target_col = find_header_columns(rows)
            for row_index, row in enumerate(rows[start_row:], start=start_row + 1):
                source = clean_cell(row[source_col] if source_col < len(row) else "")
                target = clean_target_cell(row[target_col] if target_col is not None and target_col < len(row) else "")
                if source:
                    term_rows.append(TermRow(sheet_name, row_index, sheet_name, source, target))
        return term_rows
    finally:
        workbook.close()


def read_summary_term_rows(workbook) -> list[TermRow]:
    term_rows: list[TermRow] = []
    rows = list(workbook[SUMMARY_SHEET].iter_rows(values_only=True))
    if not rows:
        return term_rows

    start_row, term_type_col, source_col, target_col = find_summary_header_columns(rows)
    for row_index, row in enumerate(rows[start_row:], start=start_row + 1):
        term_type = clean_cell(row[term_type_col] if term_type_col < len(row) else "")
        source = clean_cell(row[source_col] if source_col < len(row) else "")
        target = clean_target_cell(row[target_col] if target_col is not None and target_col < len(row) else "")
        if not source:
            continue
        if term_type not in TERM_SHEETS:
            raise ValueError(f"{SUMMARY_SHEET} 的 term_type 需要是 {STEM_SHEET} 或 {SUFFIX_SHEET}：{term_type}")
        term_rows.append(TermRow(SUMMARY_SHEET, row_index, term_type, source, target))
    return term_rows


def read_summary_terms_entries(workbook) -> dict[str, list[tuple[str, str]]]:
    entries: dict[str, list[tuple[str, str]]] = {sheet_name: [] for sheet_name in TERM_SHEETS}
    rows = list(workbook[SUMMARY_SHEET].iter_rows(values_only=True))
    if not rows:
        return entries

    start_row, term_type_col, source_col, target_col = find_summary_header_columns(rows)
    for row in rows[start_row:]:
        term_type = clean_cell(row[term_type_col] if term_type_col < len(row) else "")
        source = clean_cell(row[source_col] if source_col < len(row) else "")
        target = clean_target_cell(row[target_col] if target_col is not None and target_col < len(row) else "")
        if not source:
            continue
        if term_type not in entries:
            raise ValueError(f"{SUMMARY_SHEET} 的 term_type 需要是 {STEM_SHEET} 或 {SUFFIX_SHEET}：{term_type}")
        entries[term_type].append((source, target))
    return entries


def read_completed_terms(path: Path) -> dict[str, dict[str, str]]:
    entries = read_terms_entries(path)
    completed: dict[str, dict[str, str]] = {sheet_name: {} for sheet_name in TERM_SHEETS}

    for sheet_name in TERM_SHEETS:
        for source, target in entries.get(sheet_name, []):
            source = clean_cell(source)
            target = clean_target_cell(target)
            if not source or not target:
                continue
            existing_target = completed[sheet_name].get(source)
            if existing_target is not None and existing_target != target:
                raise ValueError(f"已翻译术语表内同一 source 有不同 target：{sheet_name} / {source}")
            completed[sheet_name][source] = target

    return completed


def ensure_named_column(sheet, names: set[str], header_value: str) -> int:
    headers = [normalize_header(cell.value) for cell in sheet[1]]
    column = find_header_column(headers, names)
    if column is not None:
        return column + 1

    if sheet.max_row == 1 and sheet.max_column == 1 and not clean_cell(sheet.cell(row=1, column=1).value):
        sheet.cell(row=1, column=1, value=header_value)
        return 1

    column = sheet.max_column + 1
    sheet.cell(row=1, column=column, value=header_value)
    return column


def ensure_source_target_columns(sheet) -> tuple[int, int]:
    source_col = ensure_named_column(sheet, {"source", "src", "源文", "原文", "源语言", "源术语"}, "source")
    target_col = ensure_named_column(sheet, {"target", "tgt", "译文", "目标", "目标语言", "目标术语"}, "target")
    return source_col, target_col


def ensure_summary_columns(sheet) -> tuple[int, int, int]:
    term_type_col = ensure_named_column(sheet, {"termtype", "类型", "术语类型"}, "term_type")
    source_col = ensure_named_column(sheet, {"source", "src", "源文", "原文", "源语言", "源术语"}, "source")
    target_col = ensure_named_column(sheet, {"target", "tgt", "译文", "目标", "目标语言", "目标术语"}, "target")
    return term_type_col, source_col, target_col


def collect_sheet_targets(sheet, source_col: int, target_col: int) -> dict[str, tuple[int, str]]:
    targets: dict[str, tuple[int, str]] = {}
    for row_index in range(2, sheet.max_row + 1):
        source = clean_cell(sheet.cell(row=row_index, column=source_col).value)
        target = clean_target_cell(sheet.cell(row=row_index, column=target_col).value)
        if source and source not in targets:
            targets[source] = (row_index, target)
    return targets


def collect_summary_targets(
    sheet,
    term_type_col: int,
    source_col: int,
    target_col: int,
) -> dict[tuple[str, str], tuple[int, str]]:
    targets: dict[tuple[str, str], tuple[int, str]] = {}
    for row_index in range(2, sheet.max_row + 1):
        term_type = clean_cell(sheet.cell(row=row_index, column=term_type_col).value)
        source = clean_cell(sheet.cell(row=row_index, column=source_col).value)
        target = clean_target_cell(sheet.cell(row=row_index, column=target_col).value)
        if term_type in TERM_SHEETS and source and (term_type, source) not in targets:
            targets[(term_type, source)] = (row_index, target)
    return targets


def format_tb_conflicts(conflicts: list[tuple[str, str, str, str]]) -> str:
    details = [
        f"{term_type} / {source}：过往 target={old_target}；本批次 target={new_target}"
        for term_type, source, old_target, new_target in conflicts
    ]
    return "过往 TB 已有不同译文，请先回到检查步骤处理冲突：" + "；".join(details)


def update_previous_tb_workbook(translated_terms_path: Path, previous_tb_path: Path) -> int:
    completed_terms = read_completed_terms(translated_terms_path)
    workbook = load_workbook(previous_tb_path)
    try:
        if SUMMARY_SHEET in workbook.sheetnames:
            updated_count = update_summary_tb_workbook(workbook, completed_terms)
        else:
            updated_count = update_split_tb_workbook(workbook, completed_terms)
        workbook.save(previous_tb_path)
        return updated_count
    finally:
        workbook.close()


def update_split_tb_workbook(workbook, completed_terms: dict[str, dict[str, str]]) -> int:
    sheet_contexts = {}
    conflicts: list[tuple[str, str, str, str]] = []

    for sheet_name in TERM_SHEETS:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        source_col, target_col = ensure_source_target_columns(sheet)
        existing = collect_sheet_targets(sheet, source_col, target_col)
        sheet_contexts[sheet_name] = (sheet, source_col, target_col, existing)

        for source, target in completed_terms[sheet_name].items():
            row = existing.get(source)
            if row is None:
                continue
            _, old_target = row
            if old_target and old_target != target:
                conflicts.append((sheet_name, source, old_target, target))

    if conflicts:
        raise ValueError(format_tb_conflicts(conflicts))

    updated_count = 0
    for sheet_name in TERM_SHEETS:
        sheet, source_col, target_col, existing = sheet_contexts[sheet_name]
        for source, target in completed_terms[sheet_name].items():
            row = existing.get(source)
            if row is None:
                row_index = sheet.max_row + 1
                sheet.cell(row=row_index, column=source_col, value=source)
                sheet.cell(row=row_index, column=target_col, value=target)
                updated_count += 1
                continue

            row_index, old_target = row
            if not old_target:
                sheet.cell(row=row_index, column=target_col, value=target)
                updated_count += 1

    return updated_count


def update_summary_tb_workbook(workbook, completed_terms: dict[str, dict[str, str]]) -> int:
    sheet = workbook[SUMMARY_SHEET]
    term_type_col, source_col, target_col = ensure_summary_columns(sheet)
    existing = collect_summary_targets(sheet, term_type_col, source_col, target_col)
    conflicts: list[tuple[str, str, str, str]] = []

    for sheet_name in TERM_SHEETS:
        for source, target in completed_terms[sheet_name].items():
            row = existing.get((sheet_name, source))
            if row is None:
                continue
            _, old_target = row
            if old_target and old_target != target:
                conflicts.append((sheet_name, source, old_target, target))

    if conflicts:
        raise ValueError(format_tb_conflicts(conflicts))

    updated_count = 0
    for sheet_name in TERM_SHEETS:
        for source, target in completed_terms[sheet_name].items():
            row = existing.get((sheet_name, source))
            if row is None:
                row_index = sheet.max_row + 1
                sheet.cell(row=row_index, column=term_type_col, value=sheet_name)
                sheet.cell(row=row_index, column=source_col, value=source)
                sheet.cell(row=row_index, column=target_col, value=target)
                updated_count += 1
                continue

            row_index, old_target = row
            if not old_target:
                sheet.cell(row=row_index, column=target_col, value=target)
                updated_count += 1

    return updated_count


def prefill_terms_workbook(batch_path: Path, previous_tb_path: Path, output_path: Path) -> None:
    batch_terms = read_terms_workbook(batch_path)
    previous_terms = read_terms_workbook(previous_tb_path)

    prefilled: dict[str, dict[str, str]] = {}
    for sheet_name in TERM_SHEETS:
        prefilled[sheet_name] = {}
        for source, current_target in batch_terms[sheet_name].items():
            target = current_target or previous_terms[sheet_name].get(source, "")
            first_value_wins_add(prefilled[sheet_name], source, target)

    write_terms_workbook(output_path, prefilled[STEM_SHEET], prefilled[SUFFIX_SHEET], include_status=True)


def check_conflicts(
    batch_terms: dict[str, dict[str, str]],
    previous_terms: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    batch_entries = {
        sheet_name: list(terms.items())
        for sheet_name, terms in batch_terms.items()
    }
    previous_entries = {
        sheet_name: list(terms.items())
        for sheet_name, terms in previous_terms.items()
    }
    return check_conflict_entries(batch_entries, previous_entries)


def check_conflict_entries(
    batch_entries: dict[str, list[tuple[str, str]]],
    previous_entries: dict[str, list[tuple[str, str]]],
) -> list[dict[str, str]]:
    conflicts: list[dict[str, str]] = []
    conflict_maps = build_conflict_maps(batch_entries, previous_entries)

    for sheet_name in TERM_SHEETS:
        by_source = conflict_maps[sheet_name]["by_source"]
        by_target = conflict_maps[sheet_name]["by_target"]

        for source, targets in by_source.items():
            if len(targets) > 1:
                conflicts.append(
                    {
                        "issue_type": "source_conflict",
                        "term_type": sheet_name,
                        "source": source,
                        "target": " | ".join(sorted(targets)),
                        "details": "Same source has multiple targets.",
                    }
                )

        for target, sources in by_target.items():
            if len(sources) > 1:
                conflicts.append(
                    {
                        "issue_type": "target_conflict",
                        "term_type": sheet_name,
                        "source": " | ".join(sorted(sources)),
                        "target": target,
                        "details": "Same target is used by multiple sources.",
                    }
                )

    return conflicts


def build_conflict_maps(
    batch_entries: dict[str, list[tuple[str, str]]],
    previous_entries: dict[str, list[tuple[str, str]]],
) -> dict[str, dict[str, dict[str, set[str]]]]:
    conflict_maps: dict[str, dict[str, dict[str, set[str]]]] = {}
    for sheet_name in TERM_SHEETS:
        by_source: dict[str, set[str]] = defaultdict(set)
        by_target: dict[str, set[str]] = defaultdict(set)

        for term_entries in (previous_entries.get(sheet_name, []), batch_entries.get(sheet_name, [])):
            for source, target in term_entries:
                source = clean_cell(source)
                target = clean_target_cell(target)
                if not source or not target:
                    continue
                by_source[source].add(target)
                by_target[target].add(source)

        conflict_maps[sheet_name] = {"by_source": by_source, "by_target": by_target}
    return conflict_maps


def targets_for_source(entries: list[tuple[str, str]], source: str) -> set[str]:
    targets: set[str] = set()
    for entry_source, entry_target in entries:
        entry_source = clean_cell(entry_source)
        entry_target = clean_target_cell(entry_target)
        if entry_source == source and entry_target:
            targets.add(entry_target)
    return targets


def build_row_issues(
    batch_rows: list[TermRow],
    batch_entries: dict[str, list[tuple[str, str]]],
    previous_entries: dict[str, list[tuple[str, str]]],
) -> dict[tuple[str, int], tuple[str, str]]:
    conflict_maps = build_conflict_maps(batch_entries, previous_entries)
    row_issues: dict[tuple[str, int], tuple[str, str]] = {}

    for term_row in batch_rows:
        if not term_row.source:
            continue
        if not term_row.target:
            row_issues[(term_row.sheet_name, term_row.row_index)] = ("missing_target", "target 为空：请补充译文后重新检查。")
            continue

        type_parts: list[str] = []
        detail_parts: list[str] = []
        by_source = conflict_maps[term_row.term_type]["by_source"]
        by_target = conflict_maps[term_row.term_type]["by_target"]

        source_targets = by_source.get(term_row.source, set())
        if len(source_targets) > 1:
            type_parts.append("source_conflict")
            batch_source_targets = targets_for_source(batch_entries.get(term_row.term_type, []), term_row.source)
            previous_source_targets = targets_for_source(previous_entries.get(term_row.term_type, []), term_row.source)
            if len(batch_source_targets) > 1:
                detail_parts.append(f"本批次内重复：同一 source 对应多个 target：{' | '.join(sorted(batch_source_targets))}")
            if previous_source_targets and term_row.target not in previous_source_targets:
                detail_parts.append(
                    "与过往术语表冲突："
                    f"过往 target：{' | '.join(sorted(previous_source_targets))}；"
                    f"本批次 target：{term_row.target}"
                )
            if not detail_parts:
                detail_parts.append(f"同一 source 对应多个 target：{' | '.join(sorted(source_targets))}")

        target_sources = by_target.get(term_row.target, set())
        if len(target_sources) > 1:
            type_parts.append("target_conflict")
            detail_parts.append(f"同一 target 被多个 source 使用：{' | '.join(sorted(target_sources))}")

        if type_parts:
            row_issues[(term_row.sheet_name, term_row.row_index)] = ("; ".join(type_parts), "\n".join(detail_parts))

    return row_issues


def ensure_issue_columns(sheet) -> tuple[int, int]:
    headers = [normalize_header(cell.value) for cell in sheet[1]]
    issue_type_col = next((index + 1 for index, name in enumerate(headers) if name in {"issuetype", "问题类型"}), None)
    issue_detail_col = next((index + 1 for index, name in enumerate(headers) if name in {"issuedetail", "问题详情"}), None)

    next_col = sheet.max_column + 1
    if issue_type_col is None:
        issue_type_col = next_col
        sheet.cell(row=1, column=issue_type_col, value="issue_type")
        next_col += 1
    if issue_detail_col is None:
        issue_detail_col = next_col
        sheet.cell(row=1, column=issue_detail_col, value="issue_detail")

    return issue_type_col, issue_detail_col


def write_check_issues(batch_path: Path, row_issues: dict[tuple[str, int], tuple[str, str]]) -> None:
    workbook = load_workbook(batch_path)
    try:
        report_sheet_names = [SUMMARY_SHEET] if SUMMARY_SHEET in workbook.sheetnames else [
            sheet_name for sheet_name in TERM_SHEETS if sheet_name in workbook.sheetnames
        ]

        for sheet_name in report_sheet_names:
            sheet = workbook[sheet_name]
            if sheet.max_row < 1:
                continue
            issue_type_col, issue_detail_col = ensure_issue_columns(sheet)
            sheet.column_dimensions[sheet.cell(row=1, column=issue_type_col).column_letter].width = 22
            sheet.column_dimensions[sheet.cell(row=1, column=issue_detail_col).column_letter].width = 70

            for row_index in range(2, sheet.max_row + 1):
                issue_type, issue_detail = row_issues.get((sheet_name, row_index), ("", ""))
                sheet.cell(row=row_index, column=issue_type_col).value = issue_type or None
                sheet.cell(row=row_index, column=issue_detail_col).value = issue_detail or None

        workbook.save(batch_path)
    finally:
        workbook.close()


def check_terms_workbook(batch_path: Path, previous_tb_path: Path) -> int:
    batch_rows = read_term_rows(batch_path)
    batch_entries: dict[str, list[tuple[str, str]]] = {sheet_name: [] for sheet_name in TERM_SHEETS}
    for term_row in batch_rows:
        batch_entries[term_row.term_type].append((term_row.source, term_row.target))

    previous_entries = read_terms_entries(previous_tb_path)
    conflicts = check_conflict_entries(batch_entries, previous_entries)
    row_issues = build_row_issues(batch_rows, batch_entries, previous_entries)
    write_check_issues(batch_path, row_issues)
    missing_target_count = sum(1 for term_row in batch_rows if term_row.source and not term_row.target)
    return len(conflicts) + missing_target_count


def compose_target(source: str, terms: dict[str, dict[str, str]]) -> str:
    source_stem, source_suffix = split_source(source)
    stem_target = terms[STEM_SHEET].get(source_stem, "")
    if not stem_target:
        return ""
    if not source_suffix:
        return stem_target

    suffix_target = terms[SUFFIX_SHEET].get(source_suffix, "")
    if not suffix_target:
        return ""
    return f"{stem_target}{suffix_target}"


def fill_demand_workbook(
    demand_path: Path,
    terms_path: Path,
    output_path: Path,
    sheet_name: str | None = None,
) -> None:
    workbook = load_workbook(demand_path)
    actual_sheet = sheet_name or find_sheet_with_source(workbook)
    if actual_sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet {actual_sheet!r} not found in {demand_path}.")

    sheet = workbook[actual_sheet]
    rows = list(sheet.iter_rows(values_only=True))
    start_row, source_col, target_col = find_header_columns(rows)
    if target_col is None:
        target_col = sheet.max_column
        sheet.cell(row=1, column=target_col + 1, value="target")

    terms = read_terms_workbook(terms_path)
    for row_index in range(start_row + 1, sheet.max_row + 1):
        source = clean_cell(sheet.cell(row=row_index, column=source_col + 1).value)
        if not source:
            continue
        target = compose_target(source, terms)
        sheet.cell(row=row_index, column=target_col + 1, value=target or None)

    workbook.save(output_path)


def default_output_path(input_path: Path, suffix: str) -> Path:
    return input_path.with_name(f"{input_path.stem}_{suffix}.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the source/target term workflow.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    split_parser = subparsers.add_parser("split", help="Split demand Excel into batch stem/suffix term sheets.")
    split_parser.add_argument("demand", type=Path)
    split_parser.add_argument("--sheet")
    split_parser.add_argument("--output", type=Path)
    split_parser.add_argument("--summary-output", type=Path)

    prefill_parser = subparsers.add_parser("prefill", help="Fill batch terms from previous TB where possible.")
    prefill_parser.add_argument("batch_terms", type=Path)
    prefill_parser.add_argument("previous_tb", type=Path)
    prefill_parser.add_argument("--output", type=Path)

    check_parser = subparsers.add_parser("check", help="Check source/target conflicts across batch terms and previous TB.")
    check_parser.add_argument("batch_terms", type=Path)
    check_parser.add_argument("previous_tb", type=Path)

    fill_parser = subparsers.add_parser("fill", help="Fill demand Excel target column from translated terms.")
    fill_parser.add_argument("demand", type=Path)
    fill_parser.add_argument("translated_terms", type=Path)
    fill_parser.add_argument("--sheet")
    fill_parser.add_argument("--output", type=Path)

    update_tb_parser = subparsers.add_parser("update-tb", help="Update previous TB in place from translated batch terms.")
    update_tb_parser.add_argument("translated_terms", type=Path)
    update_tb_parser.add_argument("previous_tb", type=Path)

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.command == "split":
        output_path = args.output or default_output_path(args.demand, "batch_terms")
        split_demand_workbook(args.demand, output_path, args.sheet, args.summary_output)
        print(f"Output: {output_path}")
        if args.summary_output:
            print(f"Summary output: {args.summary_output}")
        return

    if args.command == "prefill":
        output_path = args.output or default_output_path(args.batch_terms, "prefilled")
        prefill_terms_workbook(args.batch_terms, args.previous_tb, output_path)
        print(f"Output: {output_path}")
        return

    if args.command == "check":
        count = check_terms_workbook(args.batch_terms, args.previous_tb)
        print(f"Conflicts: {count}")
        print(f"Updated: {args.batch_terms}")
        return

    if args.command == "fill":
        output_path = args.output or default_output_path(args.demand, "filled")
        fill_demand_workbook(args.demand, args.translated_terms, output_path, args.sheet)
        print(f"Output: {output_path}")
        return

    if args.command == "update-tb":
        count = update_previous_tb_workbook(args.translated_terms, args.previous_tb)
        print(f"Updated terms: {count}")
        print(f"Updated TB: {args.previous_tb}")


if __name__ == "__main__":
    main()
