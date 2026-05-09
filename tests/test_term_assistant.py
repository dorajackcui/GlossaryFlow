import io
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path

from openpyxl import Workbook, load_workbook

from term_assistant import (
    print_header,
    run_check,
    run_fill,
    run_new_batch,
    run_prefill_batch,
    run_split_batch,
    run_update_previous_tb,
)


def write_demand(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "需求"
    sheet.append(["source"])
    sheet.append(["示例主干A·示例后缀A"])
    sheet.append(["示例主干C"])
    workbook.save(path)


def write_tb(path: Path) -> None:
    workbook = Workbook()
    stem_sheet = workbook.active
    stem_sheet.title = "主干术语表"
    stem_sheet.append(["source", "target"])
    stem_sheet.append(["示例主干A", "Target A"])
    stem_sheet.append(["示例主干C", "Target C"])
    suffix_sheet = workbook.create_sheet("后缀术语表")
    suffix_sheet.append(["source", "target"])
    suffix_sheet.append(["·示例后缀A", " : Suffix Target A"])
    workbook.save(path)


def read_rows(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()


class TermAssistantTests(unittest.TestCase):
    def test_print_header_shows_split_and_prefill_as_separate_steps(self) -> None:
        output = io.StringIO()

        with redirect_stdout(output):
            print_header()

        text = output.getvalue()
        self.assertIn("1. 开始新批次：拆分需求 Excel", text)
        self.assertIn("2. 用过往 TB 预填本批次术语表", text)
        self.assertIn("3. 我已经翻译完了：检查问题", text)
        self.assertIn("4. 检查通过了：回填需求 Excel", text)
        self.assertIn("5. 将本批次已完成术语写回过往 TB", text)
        self.assertNotIn("用 test.xlsx 跑一遍演示", text)
        self.assertNotIn("拆分需求 Excel + 用过往 TB 预填", text)

    def test_run_split_batch_creates_split_and_summary_without_prefill(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            demand = tmp_path / "需求.xlsx"
            output_root = tmp_path / "outputs"
            write_demand(demand)

            result = run_split_batch(demand, output_root)

            self.assertTrue(result.batch_terms.exists())
            self.assertTrue(result.summary_terms.exists())
            self.assertFalse((result.batch_dir / "02_本批次术语表_预填.xlsx").exists())
            self.assertIn("选择 2", result.next_step)
            self.assertEqual(
                read_rows(result.summary_terms, "汇总术语表"),
                [
                    ("term_type", "source", "target", "status"),
                    ("主干术语表", "示例主干A", None, "needs_translation"),
                    ("主干术语表", "示例主干C", None, "needs_translation"),
                    ("后缀术语表", "·示例后缀A", None, "needs_translation"),
                ],
            )

    def test_run_prefill_batch_creates_prefilled_file_from_split_terms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            demand = tmp_path / "需求.xlsx"
            tb = tmp_path / "过往TB.xlsx"
            output_root = tmp_path / "outputs"
            write_demand(demand)
            write_tb(tb)
            split = run_split_batch(demand, output_root)

            result = run_prefill_batch(split.batch_terms, tb, output_root)

            self.assertEqual(result.batch_dir, split.batch_dir)
            self.assertTrue(result.prefilled_terms.exists())
            self.assertIn("选择 3", result.next_step)
            rows = read_rows(result.prefilled_terms, "主干术语表")
            self.assertEqual(rows[1], ("示例主干A", "Target A", "matched"))

    def test_run_new_batch_creates_split_and_prefilled_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            demand = tmp_path / "需求.xlsx"
            tb = tmp_path / "过往TB.xlsx"
            output_root = tmp_path / "outputs"
            write_demand(demand)
            write_tb(tb)

            result = run_new_batch(demand, tb, output_root)

            self.assertTrue(result.batch_terms.exists())
            self.assertTrue(result.summary_terms.exists())
            self.assertTrue(result.prefilled_terms.exists())
            self.assertIn("请打开", result.next_step)
            rows = read_rows(result.prefilled_terms, "主干术语表")
            self.assertEqual(rows[1], ("示例主干A", "Target A", "matched"))

    def test_run_new_batch_defaults_outputs_next_to_demand_excel(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            demand_dir = tmp_path / "需求文件夹"
            demand_dir.mkdir()
            demand = demand_dir / "需求.xlsx"
            tb = tmp_path / "过往TB.xlsx"
            write_demand(demand)
            write_tb(tb)

            result = run_new_batch(demand, tb)

            self.assertEqual(result.batch_dir.parent.resolve(), (demand_dir / "outputs").resolve())
            self.assertTrue(result.prefilled_terms.exists())

    def test_run_check_updates_checked_terms_in_place(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            demand = tmp_path / "需求.xlsx"
            tb = tmp_path / "过往TB.xlsx"
            output_root = tmp_path / "outputs"
            write_demand(demand)
            write_tb(tb)
            batch = run_new_batch(demand, tb, output_root)

            result = run_check(batch.prefilled_terms, tb, output_root)

            self.assertEqual(result.conflict_count, 0)
            self.assertEqual(result.updated_terms, batch.prefilled_terms)
            self.assertFalse((batch.batch_dir / "03_检查报告.xlsx").exists())
            rows = read_rows(batch.prefilled_terms, "主干术语表")
            self.assertEqual(rows[0], ("source", "target", "status", "issue_type", "issue_detail"))
            self.assertIsNone(rows[1][3])
            self.assertIsNone(rows[1][4])

    def test_run_fill_creates_filled_demand(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            demand = tmp_path / "需求.xlsx"
            tb = tmp_path / "过往TB.xlsx"
            output_root = tmp_path / "outputs"
            write_demand(demand)
            write_tb(tb)
            batch = run_new_batch(demand, tb, output_root)

            result = run_fill(demand, batch.prefilled_terms, output_root)

            self.assertTrue(result.filled_demand.exists())
            rows = read_rows(result.filled_demand, "需求")
            self.assertEqual(rows[1], ("示例主干A·示例后缀A", "Target A : Suffix Target A"))

    def test_run_update_previous_tb_updates_tb_in_place(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            translated = tmp_path / "已翻译术语表.xlsx"
            tb = tmp_path / "过往TB.xlsx"
            output_root = tmp_path / "outputs"
            write_tb(tb)

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target"])
            stem_sheet.append(["示例主干A", "Target A"])
            stem_sheet.append(["示例主干B", "Target E"])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target"])
            suffix_sheet.append(["·示例后缀C", " : Suffix Target C"])
            workbook.save(translated)

            result = run_update_previous_tb(translated, tb, output_root)

            self.assertEqual(result.updated_tb, tb)
            self.assertEqual(result.updated_count, 2)
            self.assertIn("闭环完成", result.next_step)
            self.assertIn(("示例主干B", "Target E"), read_rows(tb, "主干术语表"))
            self.assertIn(("·示例后缀C", " : Suffix Target C"), read_rows(tb, "后缀术语表"))


if __name__ == "__main__":
    unittest.main()
