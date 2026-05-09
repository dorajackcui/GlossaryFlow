import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import term_workflow
from term_workflow import (
    check_conflicts,
    check_terms_workbook,
    fill_demand_workbook,
    prefill_terms_workbook,
    split_demand_workbook,
    update_previous_tb_workbook,
)


def write_rows(path: Path, sheet_name: str, rows: list[list[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def read_sheet(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()


class TermWorkflowTests(unittest.TestCase):
    def test_split_demand_workbook_writes_unique_source_only_terms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            demand = Path(tmp) / "demand.xlsx"
            output = Path(tmp) / "batch_terms.xlsx"
            write_rows(
                demand,
                "需求",
                [
                    ["source"],
                    ["示例主干A·示例后缀A"],
                    ["示例主干B·示例后缀A"],
                    ["示例主干C"],
                    ["示例主干A·示例后缀B"],
                ],
            )

            split_demand_workbook(demand, output)

            self.assertEqual(
                read_sheet(output, "主干术语表"),
                [
                    ("source", "target", "status"),
                    ("示例主干A", None, "needs_translation"),
                    ("示例主干B", None, "needs_translation"),
                    ("示例主干C", None, "needs_translation"),
                ],
            )
            self.assertEqual(
                read_sheet(output, "后缀术语表"),
                [
                    ("source", "target", "status"),
                    ("·示例后缀A", None, "needs_translation"),
                    ("·示例后缀B", None, "needs_translation"),
                ],
            )

    def test_split_demand_workbook_preserves_existing_targets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            demand = Path(tmp) / "demand.xlsx"
            output = Path(tmp) / "batch_terms.xlsx"
            write_rows(
                demand,
                "需求",
                [
                    ["source", "target"],
                    ["示例主干A·示例后缀A", "Target A : Suffix Target A"],
                    ["示例主干C", ""],
                ],
            )

            split_demand_workbook(demand, output)

            self.assertEqual(
                read_sheet(output, "主干术语表"),
                [
                    ("source", "target", "status"),
                    ("示例主干A", "Target A", "matched"),
                    ("示例主干C", None, "needs_translation"),
                ],
            )
            self.assertEqual(
                read_sheet(output, "后缀术语表"),
                [
                    ("source", "target", "status"),
                    ("·示例后缀A", " : Suffix Target A", "matched"),
                ],
            )

    def test_split_demand_workbook_splits_hyphen_source_and_target_suffixes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            demand = Path(tmp) / "demand.xlsx"
            output = Path(tmp) / "batch_terms.xlsx"
            write_rows(
                demand,
                "需求",
                [
                    ["source", "target"],
                    ["示例主干-示例后缀", "Example stem-Example suffix"],
                ],
            )

            split_demand_workbook(demand, output)

            self.assertEqual(
                read_sheet(output, "主干术语表"),
                [
                    ("source", "target", "status"),
                    ("示例主干", "Example stem", "matched"),
                ],
            )
            self.assertEqual(
                read_sheet(output, "后缀术语表"),
                [
                    ("source", "target", "status"),
                    ("-示例后缀", "-Example suffix", "matched"),
                ],
            )

    def test_split_demand_workbook_uses_first_non_empty_target_for_repeated_terms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            demand = Path(tmp) / "demand.xlsx"
            output = Path(tmp) / "batch_terms.xlsx"
            write_rows(
                demand,
                "需求",
                [
                    ["source", "target"],
                    ["示例主干A·示例后缀A", ""],
                    ["示例主干A·示例后缀A", "Target A : Suffix Target A"],
                    ["示例主干A·示例后缀A", "Target B : Suffix Target B"],
                ],
            )

            split_demand_workbook(demand, output)

            self.assertEqual(
                read_sheet(output, "主干术语表"),
                [
                    ("source", "target", "status"),
                    ("示例主干A", "Target A", "matched"),
                ],
            )
            self.assertEqual(
                read_sheet(output, "后缀术语表"),
                [
                    ("source", "target", "status"),
                    ("·示例后缀A", " : Suffix Target A", "matched"),
                ],
            )

    def test_split_demand_workbook_can_write_summary_terms_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            demand = Path(tmp) / "demand.xlsx"
            output = Path(tmp) / "batch_terms.xlsx"
            summary_output = Path(tmp) / "batch_terms_summary.xlsx"
            write_rows(
                demand,
                "需求",
                [
                    ["source", "target"],
                    ["示例主干A·示例后缀A", "Target A : Suffix Target A"],
                    ["示例主干C", ""],
                ],
            )

            split_demand_workbook(demand, output, summary_output_path=summary_output)

            self.assertEqual(
                read_sheet(summary_output, "汇总术语表"),
                [
                    ("term_type", "source", "target", "status"),
                    ("主干术语表", "示例主干A", "Target A", "matched"),
                    ("主干术语表", "示例主干C", None, "needs_translation"),
                    ("后缀术语表", "·示例后缀A", " : Suffix Target A", "matched"),
                ],
            )

    def test_prefill_terms_workbook_accepts_summary_terms_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            batch = Path(tmp) / "batch_terms_summary.xlsx"
            previous = Path(tmp) / "previous_tb.xlsx"
            output = Path(tmp) / "prefilled.xlsx"

            write_rows(
                batch,
                "汇总术语表",
                [
                    ["term_type", "source", "target", "status"],
                    ["主干术语表", "示例主干A", None, "needs_translation"],
                    ["后缀术语表", "·示例后缀A", None, "needs_translation"],
                ],
            )

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target"])
            stem_sheet.append(["示例主干A", "Target A"])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target"])
            suffix_sheet.append(["·示例后缀A", " : Suffix Target A"])
            workbook.save(previous)

            prefill_terms_workbook(batch, previous, output)

            self.assertEqual(
                read_sheet(output, "主干术语表"),
                [
                    ("source", "target", "status"),
                    ("示例主干A", "Target A", "matched"),
                ],
            )
            self.assertEqual(
                read_sheet(output, "后缀术语表"),
                [
                    ("source", "target", "status"),
                    ("·示例后缀A", " : Suffix Target A", "matched"),
                ],
            )

    def test_check_terms_workbook_accepts_summary_terms_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            batch = Path(tmp) / "batch_terms_summary.xlsx"
            previous = Path(tmp) / "previous_tb.xlsx"
            report = Path(tmp) / "report.xlsx"

            write_rows(
                batch,
                "汇总术语表",
                [
                    ["term_type", "source", "target", "status", "issue_type", "issue_detail"],
                    ["主干术语表", "示例主干A", "Target A", "matched", "old_issue", "old detail"],
                    ["主干术语表", "示例主干A", "Target B", "matched", "old_issue", "old detail"],
                    ["主干术语表", "示例主干C", "Target C", "matched", "old_issue", "old detail"],
                ],
            )
            write_rows(
                previous,
                "汇总术语表",
                [
                    ["term_type", "source", "target"],
                    ["主干术语表", "示例主干A", "Previous Target A"],
                    ["主干术语表", "示例主干D", "Target A"],
                ],
            )

            conflict_count = check_terms_workbook(batch, previous)

            self.assertFalse(report.exists())
            report_rows = read_sheet(batch, "汇总术语表")
            self.assertEqual(conflict_count, 2)
            self.assertEqual(
                report_rows[0],
                ("term_type", "source", "target", "status", "issue_type", "issue_detail"),
            )
            self.assertEqual(report_rows[1][4], "source_conflict; target_conflict")
            self.assertIn("本批次内重复", report_rows[1][5])
            self.assertIn("与过往术语表冲突", report_rows[1][5])
            self.assertIn("过往 target：Previous Target A", report_rows[1][5])
            self.assertIn("本批次 target：Target A", report_rows[1][5])
            self.assertIn("Target A", report_rows[1][5])
            self.assertIn("Target B", report_rows[1][5])
            self.assertIn("示例主干D", report_rows[1][5])
            self.assertEqual(report_rows[2][4], "source_conflict")
            self.assertIn("本批次内重复", report_rows[2][5])
            self.assertIn("与过往术语表冲突", report_rows[2][5])
            self.assertIn("本批次 target：Target B", report_rows[2][5])
            self.assertIsNone(report_rows[3][4])
            self.assertIsNone(report_rows[3][5])

    def test_check_terms_workbook_reports_missing_targets_in_place(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            batch = Path(tmp) / "batch_terms.xlsx"
            previous = Path(tmp) / "previous_tb.xlsx"

            write_rows(
                batch,
                "主干术语表",
                [
                    ["source", "target", "issue_type", "issue_detail"],
                    ["示例主干A", None, "old_issue", "old detail"],
                    ["示例主干C", "Target C", "old_issue", "old detail"],
                ],
            )
            workbook = load_workbook(batch)
            workbook.create_sheet("后缀术语表").append(["source", "target"])
            workbook.save(batch)
            workbook.close()

            workbook = Workbook()
            workbook.active.title = "主干术语表"
            workbook.active.append(["source", "target"])
            workbook.create_sheet("后缀术语表").append(["source", "target"])
            workbook.save(previous)

            issue_count = check_terms_workbook(batch, previous)

            rows = read_sheet(batch, "主干术语表")
            self.assertEqual(issue_count, 1)
            self.assertEqual(rows[1][2], "missing_target")
            self.assertIn("target 为空", rows[1][3])
            self.assertIsNone(rows[2][2])
            self.assertIsNone(rows[2][3])

    def test_check_terms_workbook_closes_read_only_workbooks_before_writing_in_place(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            batch = Path(tmp) / "batch_terms.xlsx"
            previous = Path(tmp) / "previous_tb.xlsx"
            write_rows(
                batch,
                "主干术语表",
                [
                    ["source", "target"],
                    ["示例主干A", "Target A"],
                ],
            )
            workbook = load_workbook(batch)
            workbook.create_sheet("后缀术语表").append(["source", "target"])
            workbook.save(batch)
            workbook.close()

            workbook = Workbook()
            workbook.active.title = "主干术语表"
            workbook.active.append(["source", "target"])
            workbook.create_sheet("后缀术语表").append(["source", "target"])
            workbook.save(previous)

            read_only_workbooks: list[dict[str, bool]] = []
            original_load_workbook = term_workflow.load_workbook

            def tracking_load_workbook(*args, **kwargs):
                workbook = original_load_workbook(*args, **kwargs)
                if kwargs.get("read_only"):
                    state = {"closed": False}
                    original_close = workbook.close

                    def close() -> None:
                        state["closed"] = True
                        original_close()

                    workbook.close = close
                    read_only_workbooks.append(state)
                else:
                    self.assertTrue(all(state["closed"] for state in read_only_workbooks))
                return workbook

            with patch("term_workflow.load_workbook", side_effect=tracking_load_workbook):
                check_terms_workbook(batch, previous)

    def test_fill_demand_workbook_accepts_summary_terms_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            demand = Path(tmp) / "demand.xlsx"
            terms = Path(tmp) / "translated_terms_summary.xlsx"
            output = Path(tmp) / "filled.xlsx"

            write_rows(
                demand,
                "需求",
                [
                    ["source"],
                    ["示例主干A·示例后缀A"],
                    ["示例主干C"],
                ],
            )
            write_rows(
                terms,
                "汇总术语表",
                [
                    ["term_type", "source", "target", "status"],
                    ["主干术语表", "示例主干A", "Target A", "matched"],
                    ["后缀术语表", "·示例后缀A", " : Suffix Target A", "matched"],
                    ["主干术语表", "示例主干C", "Target C", "matched"],
                ],
            )

            fill_demand_workbook(demand, terms, output)

            self.assertEqual(
                read_sheet(output, "需求"),
                [
                    ("source", "target"),
                    ("示例主干A·示例后缀A", "Target A : Suffix Target A"),
                    ("示例主干C", "Target C"),
                ],
            )

    def test_update_previous_tb_workbook_updates_two_sheet_tb_from_two_sheet_terms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            translated = Path(tmp) / "translated_terms.xlsx"
            previous = Path(tmp) / "previous_tb.xlsx"

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target", "status", "issue_type", "issue_detail"])
            stem_sheet.append(["示例主干A", "Target A", "matched", None, None])
            stem_sheet.append(["示例主干C", "Target C", "matched", None, None])
            stem_sheet.append(["空目标", None, "needs_translation", None, None])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target", "status"])
            suffix_sheet.append(["·示例后缀A", " : Suffix Target A", "matched"])
            workbook.save(translated)

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target"])
            stem_sheet.append(["示例主干A", "Target A"])
            stem_sheet.append(["示例主干C", None])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target"])
            workbook.save(previous)

            updated_count = update_previous_tb_workbook(translated, previous)

            self.assertEqual(updated_count, 2)
            self.assertEqual(
                read_sheet(previous, "主干术语表"),
                [
                    ("source", "target"),
                    ("示例主干A", "Target A"),
                    ("示例主干C", "Target C"),
                ],
            )
            self.assertEqual(
                read_sheet(previous, "后缀术语表"),
                [
                    ("source", "target"),
                    ("·示例后缀A", " : Suffix Target A"),
                ],
            )

    def test_update_previous_tb_workbook_updates_summary_tb_from_summary_terms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            translated = Path(tmp) / "translated_terms_summary.xlsx"
            previous = Path(tmp) / "previous_tb_summary.xlsx"

            write_rows(
                translated,
                "汇总术语表",
                [
                    ["term_type", "source", "target", "status"],
                    ["主干术语表", "示例主干A", "Target A", "matched"],
                    ["主干术语表", "示例主干C", "Target C", "matched"],
                    ["后缀术语表", "·示例后缀A", " : Suffix Target A", "matched"],
                ],
            )
            write_rows(
                previous,
                "汇总术语表",
                [
                    ["term_type", "source", "target"],
                    ["主干术语表", "示例主干A", "Target A"],
                    ["主干术语表", "示例主干C", None],
                ],
            )

            updated_count = update_previous_tb_workbook(translated, previous)

            self.assertEqual(updated_count, 2)
            self.assertEqual(
                read_sheet(previous, "汇总术语表"),
                [
                    ("term_type", "source", "target"),
                    ("主干术语表", "示例主干A", "Target A"),
                    ("主干术语表", "示例主干C", "Target C"),
                    ("后缀术语表", "·示例后缀A", " : Suffix Target A"),
                ],
            )

    def test_update_previous_tb_workbook_rejects_conflicts_without_saving(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            translated = Path(tmp) / "translated_terms.xlsx"
            previous = Path(tmp) / "previous_tb.xlsx"

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target"])
            stem_sheet.append(["示例主干A", "Target A"])
            stem_sheet.append(["示例主干C", "Target C"])
            workbook.create_sheet("后缀术语表").append(["source", "target"])
            workbook.save(translated)

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target"])
            stem_sheet.append(["示例主干A", "Previous Target A"])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target"])
            workbook.save(previous)

            with self.assertRaisesRegex(ValueError, "过往 TB 已有不同译文"):
                update_previous_tb_workbook(translated, previous)

            self.assertEqual(
                read_sheet(previous, "主干术语表"),
                [
                    ("source", "target"),
                    ("示例主干A", "Previous Target A"),
                ],
            )

    def test_prefill_terms_workbook_uses_previous_tb_first_value_by_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            batch = Path(tmp) / "batch_terms.xlsx"
            previous = Path(tmp) / "previous_tb.xlsx"
            output = Path(tmp) / "prefilled.xlsx"

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target", "status"])
            stem_sheet.append(["示例主干A", None, "needs_translation"])
            stem_sheet.append(["示例主干C", None, "needs_translation"])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target", "status"])
            suffix_sheet.append(["·示例后缀A", None, "needs_translation"])
            workbook.save(batch)

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target"])
            stem_sheet.append(["示例主干A", "Target A"])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target"])
            suffix_sheet.append(["·示例后缀A", " : Suffix Target A"])
            workbook.save(previous)

            prefill_terms_workbook(batch, previous, output)

            self.assertEqual(
                read_sheet(output, "主干术语表"),
                [
                    ("source", "target", "status"),
                    ("示例主干A", "Target A", "matched"),
                    ("示例主干C", None, "needs_translation"),
                ],
            )
            self.assertEqual(
                read_sheet(output, "后缀术语表"),
                [
                    ("source", "target", "status"),
                    ("·示例后缀A", " : Suffix Target A", "matched"),
                ],
            )

    def test_check_conflicts_reports_source_and_target_conflicts(self) -> None:
        conflicts = check_conflicts(
            {
                "主干术语表": {
                    "示例主干A": "Target A",
                    "示例主干D": "Target A",
                    "示例主干C": "Target D",
                },
            },
            {
                "主干术语表": {
                    "示例主干A": "Target B",
                    "示例主干B": "Target F",
                },
            },
        )

        self.assertEqual(len(conflicts), 2)
        self.assertEqual(conflicts[0]["issue_type"], "source_conflict")
        self.assertEqual(conflicts[0]["source"], "示例主干A")
        self.assertEqual(conflicts[1]["issue_type"], "target_conflict")
        self.assertEqual(conflicts[1]["target"], "Target A")

    def test_check_terms_workbook_reports_duplicate_rows_before_deduping(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            batch = Path(tmp) / "batch_terms.xlsx"
            previous = Path(tmp) / "previous_tb.xlsx"
            report = Path(tmp) / "report.xlsx"

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target", "issue_type", "issue_detail"])
            stem_sheet.append(["示例主干A", "Target A", "old_issue", "old detail"])
            stem_sheet.append(["示例主干A", "Target B", "old_issue", "old detail"])
            stem_sheet.append(["示例主干D", "Target A", "old_issue", "old detail"])
            stem_sheet.append(["示例主干C", "Target C", "old_issue", "old detail"])
            workbook.create_sheet("后缀术语表").append(["source", "target"])
            workbook.save(batch)

            workbook = Workbook()
            workbook.active.title = "主干术语表"
            workbook.active.append(["source", "target"])
            workbook.create_sheet("后缀术语表").append(["source", "target"])
            workbook.save(previous)

            conflict_count = check_terms_workbook(batch, previous)
            report_rows = read_sheet(batch, "主干术语表")

            self.assertFalse(report.exists())
            self.assertEqual(conflict_count, 2)
            self.assertEqual(report_rows[0], ("source", "target", "issue_type", "issue_detail"))
            self.assertEqual(report_rows[1][2], "source_conflict; target_conflict")
            self.assertIn("本批次内重复", report_rows[1][3])
            self.assertNotIn("与过往术语表冲突", report_rows[1][3])
            self.assertIn("Target B", report_rows[1][3])
            self.assertIn("示例主干D", report_rows[1][3])
            self.assertEqual(report_rows[2][2], "source_conflict")
            self.assertIn("本批次内重复", report_rows[2][3])
            self.assertNotIn("与过往术语表冲突", report_rows[2][3])
            self.assertIn("Target A", report_rows[2][3])
            self.assertEqual(report_rows[3][2], "target_conflict")
            self.assertIn("示例主干A", report_rows[3][3])
            self.assertIsNone(report_rows[4][2])
            self.assertIsNone(report_rows[4][3])

    def test_fill_demand_workbook_combines_stem_and_suffix_targets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            demand = Path(tmp) / "demand.xlsx"
            terms = Path(tmp) / "translated_terms.xlsx"
            output = Path(tmp) / "filled.xlsx"

            write_rows(
                demand,
                "需求",
                [
                    ["source"],
                    ["示例主干A·示例后缀A"],
                    ["示例主干C"],
                    ["未知·示例后缀A"],
                ],
            )

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target"])
            stem_sheet.append(["示例主干A", "Target A"])
            stem_sheet.append(["示例主干C", "Target C"])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target"])
            suffix_sheet.append(["·示例后缀A", " : Suffix Target A"])
            workbook.save(terms)

            fill_demand_workbook(demand, terms, output)

            self.assertEqual(
                read_sheet(output, "需求"),
                [
                    ("source", "target"),
                    ("示例主干A·示例后缀A", "Target A : Suffix Target A"),
                    ("示例主干C", "Target C"),
                    ("未知·示例后缀A", None),
                ],
            )

    def test_fill_demand_workbook_combines_hyphen_suffix_targets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            demand = Path(tmp) / "demand.xlsx"
            terms = Path(tmp) / "translated_terms.xlsx"
            output = Path(tmp) / "filled.xlsx"

            write_rows(
                demand,
                "需求",
                [
                    ["source"],
                    ["示例主干-示例后缀"],
                ],
            )

            workbook = Workbook()
            stem_sheet = workbook.active
            stem_sheet.title = "主干术语表"
            stem_sheet.append(["source", "target"])
            stem_sheet.append(["示例主干", "Example stem"])
            suffix_sheet = workbook.create_sheet("后缀术语表")
            suffix_sheet.append(["source", "target"])
            suffix_sheet.append(["-示例后缀", "-Example suffix"])
            workbook.save(terms)

            fill_demand_workbook(demand, terms, output)

            self.assertEqual(
                read_sheet(output, "需求"),
                [
                    ("source", "target"),
                    ("示例主干-示例后缀", "Example stem-Example suffix"),
                ],
            )


if __name__ == "__main__":
    unittest.main()
